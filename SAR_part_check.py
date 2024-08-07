import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
from datetime import datetime

# Updated file path
file_path = 'C:/Users/Jiraphan.Detchokul/Documents/SAR Parts tracking.xlsx'

# Read the Excel file into a DataFrame
df = pd.read_excel(file_path, engine='openpyxl')

# Convert columns 7, 8, and 9 to datetime if they are not already
df.iloc[:, 6] = pd.to_datetime(df.iloc[:, 6], errors='coerce').dt.date
df.iloc[:, 7] = pd.to_datetime(df.iloc[:, 7], errors='coerce').dt.date
df.iloc[:, 8] = pd.to_datetime(df.iloc[:, 8], errors='coerce').dt.date

# Filter rows where column 6 is blank and column 13 is blank or "NO PO"
condition1 = df.iloc[:, 5].isna()
condition2 = df.iloc[:, 12].isna() | df.iloc[:, 12].str.strip().eq('NO PO')
filtered_df_1 = df[condition1 & condition2]

# Filter rows where column 6 is blank and column 8 is less than column 9
condition3 = df.iloc[:, 5].isna() & (df.iloc[:, 8] < df.iloc[:, 7])
filtered_df_2 = df[condition3]

# Filter rows where column 6 is blank and column 9 is less than today
today = datetime.today().date()
condition4 = df.iloc[:, 5].isna() & (df.iloc[:, 8] < today)
filtered_df_3 = df[condition4]

# Combine all filtered rows for output
combined_filtered_df = pd.concat([filtered_df_1, filtered_df_2, filtered_df_3]).drop_duplicates()

# Remove columns 15 and 16
combined_filtered_df.drop(combined_filtered_df.columns[[14, 15]], axis=1, inplace=True)

# Add "Day since added" column
combined_filtered_df['Day since added'] = (pd.to_datetime(today) - pd.to_datetime(combined_filtered_df.iloc[:, 6])).dt.days

# Define the path for the new Excel file
output_file_path = os.path.join(os.path.dirname(file_path), 'filtered_rows_selected_columns.xlsx')

# Write the combined filtered rows to the new Excel file
combined_filtered_df.to_excel(output_file_path, index=False, engine='openpyxl')

# Load the workbook and select the active worksheet
wb = openpyxl.load_workbook(output_file_path)
ws = wb.active

# Define the fills for highlighting
highlight_yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
highlight_red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
highlight_orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

# Format columns 7, 8, and 9 to only show the date
date_format = 'dd-mm-yyyy'
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=9):
    for cell in row:
        cell.number_format = date_format

# Loop through the rows and apply the correct highlighting based on conditions
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    col_6_value = row[5].value
    col_7_date = row[6].value
    col_8_date = row[7].value
    col_9_date = row[8].value
    col_13_value = row[12].value
    highlight_yellow = False
    highlight_red = False
    highlight_orange = False

    if isinstance(col_7_date, datetime):
        col_7_date = col_7_date.date()
    if isinstance(col_8_date, datetime):
        col_8_date = col_8_date.date()
    if isinstance(col_9_date, datetime):
        col_9_date = col_9_date.date()

    # Highlight yellow if column 6 is blank and column 8 is less than column 9
    if col_6_value is None and col_8_date and col_9_date and col_8_date < col_9_date:
        highlight_yellow = True

    # Highlight red if column 6 is blank and column 9 is less than today
    if col_6_value is None and col_9_date and col_9_date < today:
        highlight_red = True

    # Highlight orange if column 6 is blank, column 9 is less than today, and column 8 is more than today
    if col_6_value is None and col_9_date and col_9_date < today and col_8_date and col_8_date > today:
        highlight_orange = True

    # Apply the highlights
    if highlight_orange:
        for cell in row:
            cell.fill = highlight_orange_fill
    elif highlight_red:
        for cell in row:
            cell.fill = highlight_red_fill
    elif highlight_yellow:
        for cell in row:
            cell.fill = highlight_yellow_fill

# Save the workbook
wb.save(output_file_path)

print(f'Filtered rows with selected columns and highlighted rows have been written to {output_file_path}')
