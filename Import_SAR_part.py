import pandas as pd
import openpyxl

# File paths
new_file_path = 'C:/Users/Jiraphan.Detchokul/Documents/Updated_QAD_DATA.xlsx'
existing_file_path = 'C:/Users/Jiraphan.Detchokul/Documents/SAR Parts tracking.xlsx'


# Read the new and existing Excel files into DataFrames
new_df = pd.read_excel(new_file_path, engine='openpyxl')
existing_df = pd.read_excel(existing_file_path, engine='openpyxl')

# Load the existing workbook and select the active worksheet
wb = openpyxl.load_workbook(existing_file_path)
ws = wb.active

# Compare and update the rows
for index, new_row in new_df.iterrows():
    # Check if the first two columns match in the existing DataFrame
    match = existing_df[(existing_df.iloc[:, 0] == new_row.iloc[0]) & (existing_df.iloc[:, 1] == new_row.iloc[1])]
    
    if not match.empty:
        # Get the index of the matching row in the existing DataFrame
        match_index = match.index[0]
        
        # Check if column 6 contains "Picked" or "Not tracked"
        col_6_value = existing_df.iloc[match_index, 5]
        if col_6_value not in ["Picked", "Not tracked"]:
            # Update the rest of the row in the existing DataFrame with the new DataFrame's row
            for col in range(2, len(new_row)):
                ws.cell(row=match_index + 2, column=col + 1, value=new_row.iloc[col])
    else:
        # Add the new row to the bottom of the existing sheet
        ws.append(new_row.tolist())

# Format columns 7, 8, and 9 to only show the date in dd-mm-yyyy format
date_format = 'dd-mm-yyyy'
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=9):
    for cell in row:
        if cell.value is not None:
            cell.number_format = date_format

# Save the workbook with the updated data and formatting
wb.save(existing_file_path)

print(f'The existing data has been updated and saved in {existing_file_path}')
